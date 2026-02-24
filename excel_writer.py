"""엑셀 파일 저장 모듈."""

import os

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

import config


def _format_억(value):
    """금액을 억원 단위로 변환 (콤마 포맷)."""
    if not isinstance(value, (int, float)) or pd.isna(value):
        return ""
    return f"{round(float(value) / 1e8):,}"


def _format_comma(value):
    """숫자에 3자리 콤마 추가."""
    if not isinstance(value, (int, float)) or pd.isna(value):
        return ""
    if isinstance(value, float) and value == int(value):
        return f"{int(value):,}"
    return f"{value:,}"


def save_to_excel(df: pd.DataFrame, date_str: str):
    """수급 데이터를 엑셀 파일로 저장.

    Args:
        df: 전 종목 수급 DataFrame
        date_str: 날짜 문자열 (YYYYMMDD)
    """
    os.makedirs(config.DATA_DIR, exist_ok=True)

    if df.empty:
        print("[Excel] 저장할 데이터가 없습니다.")
        return

    file_path = os.path.join(config.DATA_DIR, f"수급_{date_str}.xlsx")

    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        unit_note = "(단위: 억원)"

        # 1) 전체 시트
        _write_sheet(writer, df, "전체",
                     f"투자자별 수급 현황 — {date_str} {unit_note}",
                     config.COLUMN_ORDER)

        # 2) 랭킹 시트 (외국인/기관/개인 TOP50)
        for investor, sheet_name in config.RANKING_INVESTORS.items():
            if investor not in df.columns:
                continue
            top_df = df.nlargest(50, investor).copy()
            cols = config.RANKING_COLUMN_ORDER + [investor]
            cols = [c for c in cols if c in top_df.columns]
            _write_sheet(writer, top_df, sheet_name,
                         f"{investor} 순매수 TOP 50 — {date_str} {unit_note}",
                         cols)

    # 스타일 적용
    book = load_workbook(file_path)
    for sheet_name in book.sheetnames:
        _apply_styles(book, sheet_name)
    book.save(file_path)

    print(f"[Excel] 저장 완료: {file_path}")


def _write_sheet(writer, df: pd.DataFrame, sheet_name: str, title: str,
                 columns: list):
    """DataFrame을 시트에 기록 (1행 타이틀, 2행 헤더, 3행부터 데이터)."""
    cols = [c for c in columns if c in df.columns]
    out_df = df[cols].copy()

    # 금액 컬럼 → 억원 단위 (콤마)
    money_cols = [c for c in cols if c in config.INVESTORS or c in ("시가총액", "거래대금")]
    for col in money_cols:
        if col in out_df.columns:
            out_df[col] = out_df[col].apply(_format_억)

    # 종가, 거래량 → 콤마만
    for col in ["종가", "거래량"]:
        if col in out_df.columns:
            out_df[col] = out_df[col].apply(_format_comma)

    # 티커를 문자열로 보장
    if "티커" in out_df.columns:
        out_df["티커"] = out_df["티커"].astype(str).str.zfill(6)

    out_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)

    # 타이틀 기록
    ws = writer.sheets[sheet_name]
    ws.cell(row=1, column=1, value=title)

    # 티커 컬럼을 텍스트 포맷으로 설정 (선행 0 보존)
    if "티커" in list(out_df.columns):
        ticker_col_idx = list(out_df.columns).index("티커") + 1
        for row_idx in range(3, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=ticker_col_idx)
            cell.number_format = "@"


def _apply_styles(book, sheet_name: str):
    """엑셀 시트에 스타일 적용."""
    ws = book[sheet_name]
    num_cols = ws.max_column

    # 타이틀 행 (1행)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    title_cell = ws.cell(row=1, column=1)
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")

    # 헤더 스타일 (2행)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4",
                              fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)

    header_names = []
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=2, column=col_idx)
        header_names.append(cell.value or "")
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # 투자자 컬럼 인덱스 (조건부 서식용)
    investor_col_indices = []
    for col_idx, name in enumerate(header_names, 1):
        if name in config.INVESTORS:
            investor_col_indices.append(col_idx)

    # 데이터 영역 스타일
    blue_font = Font(color="0000CC")
    red_font = Font(color="CC0000")

    for row in ws.iter_rows(min_row=3, min_col=1, max_col=num_cols):
        for cell in row:
            # 숫자 정렬
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal="right")

            # 투자자 컬럼 조건부 서식 (양수=파란, 음수=빨간)
            if cell.column in investor_col_indices and cell.value:
                val_str = str(cell.value)
                if val_str.startswith("-"):
                    cell.font = red_font
                elif val_str not in ("0", ""):
                    cell.font = blue_font

    # 등락률 컬럼 색상
    for col_idx, name in enumerate(header_names, 1):
        if name == "등락률":
            for row_idx in range(3, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if isinstance(cell.value, (int, float)):
                    if cell.value > 0:
                        cell.font = blue_font
                    elif cell.value < 0:
                        cell.font = red_font

    # 열 너비 자동 조정
    for col_idx in range(1, num_cols + 1):
        max_len = len(str(header_names[col_idx - 1])) if col_idx <= len(header_names) else 5
        for row in ws.iter_rows(min_row=3, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 25)

    # 자동 필터
    ws.auto_filter.ref = f"A2:{get_column_letter(num_cols)}{ws.max_row}"

    # 틀 고정 (A3)
    ws.freeze_panes = "A3"
